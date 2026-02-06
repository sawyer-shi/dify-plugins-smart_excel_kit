import pandas as pd
import numpy as np
import re
import os
import traceback
from collections.abc import Generator
from typing import Any

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage
from dify_plugin.entities.model.message import UserPromptMessage, TextPromptMessageContent
from tools.utils import ExcelProcessor

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelManipulatorTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        llm_model = tool_parameters.get('model_config')
        file_obj = tool_parameters.get('upload_file')
        user_prompt = tool_parameters.get('prompt')
        sheet_number = tool_parameters.get('sheet_number', 1)
        output_filename = tool_parameters.get('output_filename')

        if not isinstance(llm_model, dict): yield self.create_text_message("Error: model_config invalid."); return
        if not file_obj: yield self.create_text_message("Error: No file uploaded."); return
        if not isinstance(sheet_number, int) or sheet_number <= 0: yield self.create_text_message("Error: sheet_number must be greater than 0."); return

        # 1. 加载文件
        # 注意：这里我们获取了 df (指定sheet的数据) 和 wb (整个工作簿对象)
        df, wb, is_xlsx, origin_name, path_in, path_out = ExcelProcessor.load_file_with_copy(file_obj, sheet_number)

        try:
            # === 环境准备：确保是 Excel 格式 ===
            # 如果是 CSV 转 Excel，或者 wb 为空，初始化一个新的
            if not is_xlsx or wb is None:
                wb = Workbook()
                ws = wb.active
                ws.title = "Sheet1"
                # 将 CSV 数据写入
                if df is not None:
                    for r in dataframe_to_rows(df, index=False, header=True):
                        ws.append(r)
                
                # 修正路径后缀
                base_dir = os.path.dirname(path_out)
                base_name = os.path.splitext(os.path.basename(path_out))[0]
                new_path_out = os.path.join(base_dir, base_name + ".xlsx")
                if os.path.exists(path_out):
                    try: os.remove(path_out)
                    except: pass
                path_out = new_path_out
                # 标记现在已经是 Excel 环境
                is_xlsx = True
                ws_target = ws
            else:
                # 确保 Sheet 存在
                if sheet_number > len(wb.worksheets):
                    yield self.create_text_message(f"Error: Sheet index {sheet_number} is out of range. File has {len(wb.worksheets)} sheets.")
                    return
                ws_target = wb.worksheets[sheet_number - 1]

            # === Step 1: 数据画像 (Data Profiling) ===
            # 用于让 LLM 理解当前数据结构
            try:
                # 简单预览
                preview_str = df.head(5).to_string(index=False)
                # 列类型描述
                col_info = []
                for col in df.columns:
                    col_info.append(f"'{col}': {df[col].dtype}")
                col_info_str = ", ".join(col_info)
            except:
                preview_str = "Data preview unavailable"
                col_info_str = "Unknown columns"

            # === Step 2: 构造 Prompt (Code Generation) ===
            system_instruction = f"""
            You are an expert Python Pandas Data Engineer.
            
            === DATASET INFO (Sheet {sheet_number}) ===
            Columns: {col_info_str}
            Preview:
            {preview_str}
            
            === USER INSTRUCTION ===
            "{user_prompt}"

            === TASK ===
            Write a Python function named `process_data(df)` that receives the pandas DataFrame and returns the MODIFIED DataFrame.
            
            === RULES ===
            1. ONLY return the Python code block (inside ```python ... ```).
            2. The function signature MUST be `def process_data(df):`. 
            3. Do NOT load the file again. Work on the passed `df` arg.
            4. Import required libs (e.g. `import pandas as pd`, `import numpy as np`) inside the code.
            5. Ensure the returned object is a DataFrame.

            Example:
            ```python
            import pandas as pd
            def process_data(df):
                # Logic to drop rows or add columns
                df['Total'] = df['Qty'] * df['Price']
                return df
            ```
            """

            # === Step 3: 调用 LLM ===
            try:
                if hasattr(self, 'invoke_model'):
                    response = self.invoke_model(model=llm_model, messages=[UserPromptMessage(content=[TextPromptMessageContent(type='text', data=system_instruction)])])
                    llm_response = response.message.content
                elif hasattr(self, 'session'):
                    llm_service = self.session.model.llm
                    response = llm_service.invoke(model_config=llm_model, prompt_messages=[UserPromptMessage(content=[TextPromptMessageContent(type='text', data=system_instruction)])], stream=False)
                    llm_response = response.message.content
                else: 
                     raise AttributeError("No invoke interface")
            except Exception as e:
                yield self.create_text_message(f"AI Reasoning Failed: {str(e)}")
                return

            # === Step 4: 提取并执行代码 ===
            code_match = re.search(r'```python(.*?)```', llm_response, re.DOTALL)
            code_to_run = code_match.group(1).strip() if code_match else llm_response.replace('```', '')

            # 沙箱变量
            local_scope = {}
            global_scope = {'pd': pd, 'np': np} # 限制只能用 pd 和 np

            try:
                # 1. 定义函数
                exec(code_to_run, global_scope, local_scope)
                
                if 'process_data' not in local_scope:
                    raise ValueError("AI did not define 'process_data(df)' function.")
                
                # 2. 执行数据处理
                process_func = local_scope['process_data']
                # 传入副本，防止意外
                new_df = process_func(df.copy())

                if not isinstance(new_df, pd.DataFrame):
                    raise ValueError("Result is not a pandas DataFrame.")

            except Exception as code_error:
                error_details = traceback.format_exc()
                yield self.create_text_message(f"Script Execution Error:\n{error_details}\n\nCode Generated:\n{code_to_run}")
                return

            # === Step 5: 将修改后的 DF 写回 Excel 的指定 Sheet ===
            # 策略：清空当前 sheet 的内容，然后填入新数据。
            # 这样做的好处是保留了 Excel 文件中【其他 Sheet】的内容不变。
            
            # 1. 清空旧数据 (保留 Sheet 对象，不清空样式可能很难，这里主要清空值)
            # max_row 可能很大，delete_rows 比较彻底
            if ws_target.max_row > 0:
                ws_target.delete_rows(1, ws_target.max_row)
            
            # 2. 写入新数据
            # header=True (第一行写列名), index=False (不写索引)
            for r in dataframe_to_rows(new_df, index=False, header=True):
                ws_target.append(r)

            # === Step 6: 准备输出文件 ===
            # 使用统一的保存函数处理文件名
            data, fname = ExcelProcessor.save_output_file(wb, path_out, origin_name, output_filename)
            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            yield self.create_text_message(f"Processing complete on Sheet {sheet_number}.\n\nAI Logic applied:\n```python\n{code_to_run}\n```")
            yield self.create_blob_message(blob=data, meta={'mime_type': mime_type, 'save_as': fname, 'filename': fname})

        finally:
            ExcelProcessor.clean_paths([path_in, path_out])