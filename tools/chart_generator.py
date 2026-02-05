import json
import re
import os
from collections.abc import Generator
from typing import Any

from dify_plugin import Tool
from dify_plugin.entities.tool import ToolInvokeMessage
from dify_plugin.entities.model.message import UserPromptMessage, TextPromptMessageContent
from tools.utils import ExcelProcessor

# === 引入全量图表支持 ===
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import (
    AreaChart, AreaChart3D,
    BarChart, BarChart3D,
    BubbleChart,
    DoughnutChart,
    LineChart, LineChart3D,
    PieChart, PieChart3D,
    RadarChart,
    ScatterChart,
    SurfaceChart, SurfaceChart3D,
    Reference, Series
)

class ChartGeneratorTool(Tool):
    def _invoke(self, tool_parameters: dict[str, Any]) -> Generator[ToolInvokeMessage]:
        llm_model = tool_parameters.get('model_config')
        file_obj = tool_parameters.get('upload_file')
        user_prompt = tool_parameters.get('prompt')
        sheet_number = tool_parameters.get('sheet_number', 1)

        if not isinstance(llm_model, dict): yield self.create_text_message("Error: model_config invalid."); return
        if not file_obj: yield self.create_text_message("Error: No file uploaded."); return
        if not isinstance(sheet_number, int) or sheet_number <= 0: yield self.create_text_message("Error: sheet_number must be greater than 0."); return

        # 1. 加载文件
        df, wb, is_xlsx, origin_name, path_in, path_out = ExcelProcessor.load_file_with_copy(file_obj, sheet_number)
        
        try:
            # === 新增逻辑：处理 CSV 文件 ===
            # 如果是 CSV，df 有数据，但 wb 是 None。我们需要手动创建一个 wb 并把数据填进去
            if not is_xlsx or wb is None:
                # 1. 创建全新的 Excel 对象
                wb = Workbook()
                ws = wb.active
                
                # 2. 将 Pandas DataFrame 写入 Excel (使用 openpyxl 工具函数)
                # index=False 表示不写入 pandas 的索引列，header=True 表示写入表头
                for r in dataframe_to_rows(df, index=False, header=True):
                    ws.append(r)
                
                # 3. 强制修改输出路径后缀（因为不能把 Excel 二进制存进 .csv 文件）
                base_dir = os.path.dirname(path_out)
                base_name = os.path.splitext(os.path.basename(path_out))[0]
                new_path_out = os.path.join(base_dir, base_name + ".xlsx")
                
                # 删除旧的 csv 占位文件，准备保存为 xlsx
                if os.path.exists(path_out):
                    try: os.remove(path_out)
                    except: pass
                path_out = new_path_out
                
                # 4. 修改下载文件名后缀
                base_origin = os.path.splitext(origin_name)[0]
                origin_name = base_origin + ".xlsx"

                # 标记现在已经是 Excel 环境了
                is_xlsx = True

            # === 以下逻辑保持不变 ===
            if wb and sheet_number <= len(wb.worksheets):
                ws = wb.worksheets[sheet_number - 1]
            else:
                ws = wb.active if wb else None
            max_row = ws.max_row if ws else len(df)
            max_col = ws.max_column if ws else len(df.columns)

            # 数据预览
            try:
                data_preview = df.head(3).to_markdown(index=False)
                columns = df.columns.tolist()
            except Exception:
                data_preview = "Unable to read dataframe preview."
                columns = []

            # 定义图表类型描述
            chart_types_desc = """
            - 'column': Vertical Bar Chart (Clustered)
            - 'column_stacked': Vertical Bar Chart (Stacked)
            - 'bar': Horizontal Bar Chart
            - 'line': Line Chart
            - 'pie': Pie Chart
            - 'doughnut': Doughnut Chart
            - 'radar': Radar Chart
            - 'scatter': Scatter Chart (XY)
            - 'bubble': Bubble Chart
            """

            # System Prompt
            system_instruction = f"""
            You are an Excel Data Visualization Expert.
            User Request: "{user_prompt}"
            
            Data Preview:
            {data_preview}
            
            Columns: {columns}
            Total Rows: {max_row}

            TASK: Return a JSON configuration to create the chart.
            Supported 'chart_type': {chart_types_desc}

            RETURN JSON ONLY:
            {{
                "chart_type": "chart type string",
                "title": "Chart Title",
                "x_axis_col": "Column letter for X-axis (e.g., 'A')",
                "y_axis_cols": ["Column letter(s) for Y-axis", "e.g., 'B'"],
                "cell_position": "Top-left cell e.g., 'E2'"
            }}
            Rules: 
            1. 'scatter'/'bubble' use numerical X. Others use Category X.
            2. 'cell_position' should be to the right of data (e.g. column {chr(min(max_col + 66, 90))}).
            """

            # 调用 LLM
            try:
                if hasattr(self, 'invoke_model'):
                    response = self.invoke_model(model=llm_model, messages=[UserPromptMessage(content=[TextPromptMessageContent(type='text', data=system_instruction)])])
                    llm_response = response.message.content
                elif hasattr(self, 'session'):
                    llm_service = self.session.model.llm
                    response = llm_service.invoke(model_config=llm_model, prompt_messages=[UserPromptMessage(content=[TextPromptMessageContent(type='text', data=system_instruction)])], stream=False)
                    llm_response = response.message.content
                else: 
                     raise AttributeError("No invoke interface")
            except Exception as e:
                yield self.create_text_message(f"LLM Config Generation Failed: {str(e)}")
                return

            # 解析 JSON
            try:
                json_str = re.sub(r'```json\s*', '', llm_response)
                json_str = re.sub(r'```', '', json_str)
                config = json.loads(json_str.strip())
            except Exception:
                yield self.create_text_message(f"Failed to parse LLM Response: {llm_response}")
                return

            ctype = config.get('chart_type', 'column').lower()
            title = config.get('title', 'Chart')
            x_letter = config.get('x_axis_col', 'A').upper()
            y_letters = config.get('y_axis_cols', [])
            
            # 动态计算插入位置：在数据最右侧再隔开一列
            default_col_char = chr(min(max_col + 1 + 65, 90)) # 简单的 A-Z 映射，超过Z通过 LLM 指定或后续优化
            pos = config.get('cell_position', f"{default_col_char}2")

            def _c2i(letter):
                idx = 0
                for char in letter: idx = idx * 26 + (ord(char) - ord('A'))
                return idx + 1

            try:
                chart = None
                # 图表工厂
                if ctype == 'column': chart = BarChart(); chart.type = "col"; chart.style = 10
                elif ctype == 'column_stacked': chart = BarChart(); chart.type = "col"; chart.grouping = "stacked"; chart.overlap = 100
                elif ctype == 'bar': chart = BarChart(); chart.type = "bar"; chart.style = 10
                elif ctype == 'line': chart = LineChart(); chart.style = 12
                elif ctype == 'line_3d': chart = LineChart3D()
                elif ctype == 'pie': chart = PieChart()
                elif ctype == 'doughnut': chart = DoughnutChart()
                elif ctype == 'area': chart = AreaChart()
                elif ctype == 'radar': chart = RadarChart()
                elif ctype == 'scatter': chart = ScatterChart(); chart.style = 13
                elif ctype == 'bubble': chart = BubbleChart()
                elif ctype == 'surface': chart = SurfaceChart3D()
                else: chart = BarChart()

                chart.title = title
                
                x_col_idx = _c2i(x_letter)
                # Scatter/Bubble 使用数值 X 轴，其他使用 Category
                is_xy = ctype in ['scatter', 'bubble']
                
                # 定义 X 数据引用
                x_data_ref = Reference(ws, min_col=x_col_idx, min_row=2, max_row=max_row)

                if not is_xy:
                    chart.set_categories(x_data_ref)

                for y_letter in y_letters:
                    y_idx = _c2i(y_letter)
                    y_data_ref = Reference(ws, min_col=y_idx, min_row=1, max_row=max_row)
                    
                    if is_xy:
                        series = Series(y_data_ref, xvalues=x_data_ref, title_from_data=True)
                        chart.series.append(series)
                    else:
                        chart.add_data(y_data_ref, titles_from_data=True)

                ws.add_chart(chart, pos)

            except Exception as e:
                yield self.create_text_message(f"Error drawing chart: {str(e)}")
                return

            # 保存并清理
            # 注意：保存的时候，如果之前是 CSV，这里强制保存为 input.xlsx 样式
            if wb:
                wb.save(path_out) 
            
            with open(path_out, 'rb') as f:
                data = f.read()

            mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            
            # 如果原名是 data.csv，现在变成 smart_data.xlsx
            yield self.create_text_message(f"Chart '{title}' created. file converted to Excel format.")
            yield self.create_blob_message(blob=data, meta={'mime_type': mime_type, 'save_as': origin_name})

        finally:
            # 清理时要小心，可能路径已经在上面被我们改过了
            ExcelProcessor.clean_paths([path_in, path_out])