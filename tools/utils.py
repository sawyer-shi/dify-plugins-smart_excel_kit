import io
import re
import os
import requests
import base64
import tempfile
import shutil
import mimetypes
import posixpath
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from typing import List, Dict, Tuple, Any, Optional
from openpyxl import load_workbook

class ExcelProcessor:
    @staticmethod
    def load_file_with_copy(file_obj, sheet_number: int = 1) -> Tuple[pd.DataFrame, Any, bool, str, str, str]:
        """
        核心加载逻辑：
        Args:
            file_obj: 文件对象
            sheet_number: sheet页码（从1开始）
        Returns: (df, wb, is_xlsx, best_filename, input_path, output_path)
        """
        content = file_obj.blob
        
        # 1. 解析文件名
        candidates = []
        seen = set()
        common_attrs = ['original_filename', 'upload_filename', 'filename', 'name']
        for attr in common_attrs:
            if hasattr(file_obj, attr):
                val = getattr(file_obj, attr)
                if isinstance(val, str) and val not in seen:
                    candidates.append(val)
                    seen.add(val)
        try:
            if hasattr(file_obj, '__dict__'):
                val = file_obj.__dict__.get('filename')
                if isinstance(val, str) and val not in seen:
                    candidates.append(val)
                    seen.add(val)
        except: pass

        uuid_pattern = re.compile(r'^[a-fA-F0-9-]{32,36}\.(xlsx|csv|xls)$')
        best_filename = None
        for name in candidates:
            base_name = os.path.basename(name)
            if not any(base_name.lower().endswith(ext) for ext in ['.xlsx', '.csv', '.xls']): continue
            if uuid_pattern.match(base_name): continue
            best_filename = base_name; break
        
        if not best_filename: best_filename = "output.xlsx"
        _, ext = os.path.splitext(best_filename)
        ext = ext.lower()

        # 2. 创建物理文件
        fd_in, temp_input_path = tempfile.mkstemp(suffix=ext)
        with os.fdopen(fd_in, 'wb') as f:
            f.write(content)
            
        # Output: 副本
        fd_out, temp_output_path = tempfile.mkstemp(suffix=ext)
        os.close(fd_out)
        shutil.copy2(temp_input_path, temp_output_path)
        
        wb = None
        is_xlsx = False
        df = None

        try:
            if ext == '.csv':
                try: df = pd.read_csv(temp_input_path, encoding='utf-8')
                except UnicodeDecodeError: df = pd.read_csv(temp_input_path, encoding='gbk')
                df = df.fillna("")
            else:
                is_xlsx = True
                # header=0 意味着第一行是表头
                df = pd.read_excel(temp_input_path, header=0, sheet_name=sheet_number - 1)
                df = df.fillna("")
                
                if ext in ['.xlsx', '.xlsm']:
                    try: 
                        # data_only=False 保证读到公式本身而不是结果，便于写入保留
                        wb = load_workbook(temp_output_path)
                    except Exception as e: 
                        print(f"Workbook load error: {e}")
                        wb = None
                
        except Exception as e:
            ExcelProcessor.clean_paths([temp_input_path, temp_output_path])
            raise e

        return df, wb, is_xlsx, best_filename, temp_input_path, temp_output_path

    @staticmethod
    def load_file(file_obj):
        """兼容旧调用的适配器，防止 AttributeError"""
        df, wb, is_xlsx, best_filename, p_in, p_out = ExcelProcessor.load_file_with_copy(file_obj)
        # 旧接口返回5个参数，其中最后一个通常是 temp_path。
        # 我们返回 output_path 作为 temp_path
        # 注意：使用此旧接口将无法手动清理 input_path，可能会有极其微小的临时文件残留，建议更新调用方
        return df, wb, is_xlsx, best_filename, p_out

    @staticmethod
    def extract_image_map(temp_file_path: str, sheet_number: int = 1) -> Dict[Tuple[int, int], List[str]]:
        """
        强力版图片提取：兼容多种 Anchor 格式
        """
        image_map = {}
        if not temp_file_path or not os.path.exists(temp_file_path): return image_map
        if not temp_file_path.lower().endswith('.xlsx'): return image_map # 仅支持 xlsx

        wb_temp = None
        try:
            wb_temp = load_workbook(temp_file_path, data_only=False)
            if sheet_number <= len(wb_temp.worksheets):
                ws = wb_temp.worksheets[sheet_number - 1]
            else:
                ws = wb_temp.active
            
            # 尝试获取所有可能的图片列表
            images = getattr(ws, '_images', []) or getattr(ws, 'images', [])
            
            for img in images:
                try:
                    # === 核心：坐标确立 ===
                    anchor_row = None
                    anchor_col = None

                    # 策略1: 尝试访问 .anchor._from (OneCellAnchor / TwoCellAnchor)
                    if hasattr(img, 'anchor'):
                        a = img.anchor
                        # 兼容不同版本的属性名
                        marker = getattr(a, '_from', None) or getattr(a, 'from', None)
                        
                        if marker:
                            anchor_col = marker.col
                            anchor_row = marker.row
                    
                    # 如果策略1失败，且没有任何坐标信息，则跳过
                    if anchor_row is None or anchor_col is None:
                        continue

                    # === 坐标系转换 ===
                    # OpenPyxl Row 0 = Excel Row 1 (Header)
                    # OpenPyxl Row 1 = Excel Row 2 (Data Row 0)
                    # Pandas Index = OpenPyxl Row - 1
                    
                    pd_row_idx = anchor_row - 1
                    
                    if pd_row_idx < 0: continue # 图片在表头或更上面

                    # === 图片数据提取 ===
                    img_data = None
                    
                    # 方式 A: ref (通常是 BytesIO)
                    if hasattr(img, 'ref') and hasattr(img.ref, 'read'):
                        img.ref.seek(0)
                        img_data = img.ref.read()
                    # 方式 B: fp (文件路劲或流)
                    elif hasattr(img, 'fp') and hasattr(img.fp, 'read'):
                        img.fp.seek(0)
                        img_data = img.fp.read()
                    # 方式 C: _data 闭包
                    elif hasattr(img, '_data'):
                        try: img_data = img._data()
                        except: pass
                    
                    if not img_data: continue

                    # === 格式与 Base64 ===
                    fmt = "png"
                    if hasattr(img, 'format') and img.format:
                        fmt = img.format.lower()
                    
                    # 过滤不支持的格式
                    if fmt in ['emf', 'wmf', 'vml']: continue
                    if fmt not in ['png', 'jpg', 'jpeg', 'bmp', 'gif', 'webp']: fmt = 'png'

                    b64 = base64.b64encode(img_data).decode('utf-8')
                    data_uri = f"data:image/{fmt};base64,{b64}"
                    
                    key = (pd_row_idx, anchor_col)
                    if key not in image_map: image_map[key] = []
                    image_map[key].append(data_uri)

                except Exception as e:
                    print(f"Warning: Single image extract failed: {e}")
                    continue

        except Exception as e:
            print(f"Extract image map failed: {e}")

        finally:
            try:
                drawing_image_map = ExcelProcessor._extract_drawing_images_map(temp_file_path, sheet_number)
                for key, values in drawing_image_map.items():
                    if key not in image_map: image_map[key] = []
                    image_map[key].extend(values)
            except Exception as e:
                print(f"Extract drawing images failed: {e}")

            try:
                cell_image_map = ExcelProcessor._extract_cell_images_map(temp_file_path, sheet_number)
                for key, values in cell_image_map.items():
                    if key not in image_map: image_map[key] = []
                    image_map[key].extend(values)
            except Exception as e:
                print(f"Extract cell images failed: {e}")

            if wb_temp: wb_temp.close()
            
        return image_map

    @staticmethod
    def _extract_cell_images_map(temp_file_path: str, sheet_number: int) -> Dict[Tuple[int, int], List[str]]:
        image_map: Dict[Tuple[int, int], List[str]] = {}
        if not zipfile.is_zipfile(temp_file_path):
            return image_map

        with zipfile.ZipFile(temp_file_path, 'r') as zf:
            sheet_path = f"xl/worksheets/sheet{sheet_number}.xml"
            if sheet_path not in zf.namelist():
                return image_map

            sheet_xml = zf.read(sheet_path)
            try:
                sheet_root = ET.fromstring(sheet_xml)
            except Exception:
                return image_map

            sheet_parent_map = {c: p for p in sheet_root.iter() for c in p}
            sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_number}.xml.rels"
            sheet_rels = ExcelProcessor._read_relationships(zf, sheet_rels_path)

            # 1) Direct image relationships referenced from sheet
            for elem in sheet_root.iter():
                rid = ExcelProcessor._find_rid(elem)
                if not rid or rid not in sheet_rels:
                    continue

                target = sheet_rels.get(rid)
                if not target:
                    continue

                target_path = ExcelProcessor._resolve_target(sheet_rels_path, target)
                if not target_path:
                    continue

                if not target_path.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.webp')):
                    continue

                pd_row_idx, pd_col_idx = ExcelProcessor._find_cell_location(elem, sheet_parent_map)
                if pd_row_idx is None or pd_col_idx is None:
                    continue

                data_uri = ExcelProcessor._read_image_as_data_uri(zf, target_path)
                if not data_uri:
                    continue

                key = (pd_row_idx, pd_col_idx)
                if key not in image_map:
                    image_map[key] = []
                image_map[key].append(data_uri)

            # 2) Cell images part referenced from sheet rels
            for rid, target in sheet_rels.items():
                if 'cellimage' not in target.lower() and 'cellimages' not in target.lower():
                    continue
                target_path = ExcelProcessor._resolve_target(sheet_rels_path, target)
                if not target_path or target_path not in zf.namelist():
                    continue

                try:
                    cell_root = ET.fromstring(zf.read(target_path))
                except Exception:
                    continue

                cell_parent_map = {c: p for p in cell_root.iter() for c in p}
                cell_rels_path = ExcelProcessor._rels_for_part(target_path)
                cell_rels = ExcelProcessor._read_relationships(zf, cell_rels_path)

                for elem in cell_root.iter():
                    embed_rid = ExcelProcessor._find_blip_embed(elem) or ExcelProcessor._find_rid(elem)
                    if not embed_rid or embed_rid not in cell_rels:
                        continue

                    img_target = cell_rels.get(embed_rid)
                    if not img_target:
                        continue
                    img_path = ExcelProcessor._resolve_target(cell_rels_path, img_target)
                    if not img_path:
                        continue
                    data_uri = ExcelProcessor._read_image_as_data_uri(zf, img_path)
                    if not data_uri:
                        continue

                    pd_row_idx, pd_col_idx = ExcelProcessor._find_cell_location(elem, cell_parent_map)
                    if pd_row_idx is None or pd_col_idx is None:
                        continue

                    key = (pd_row_idx, pd_col_idx)
                    if key not in image_map:
                        image_map[key] = []
                    image_map[key].append(data_uri)

            return image_map

    @staticmethod
    def _resolve_images_from_sheet_rid(zf, sheet_rels_path: str, sheet_rels: Dict[str, str], rid: str,
                                       row_idx: Optional[int], col_idx: Optional[int]) -> List[Tuple[int, int, str]]:
        results: List[Tuple[int, int, str]] = []
        target = sheet_rels.get(rid)
        if not target:
            return results

        target_path = ExcelProcessor._resolve_target(sheet_rels_path, target)
        if not target_path or target_path not in zf.namelist():
            return results

        if target_path.lower().endswith('.xml'):
            # cellimage xml: contains image references and position
            xml_bytes = zf.read(target_path)
            try:
                root = ET.fromstring(xml_bytes)
            except Exception:
                return results

            cell_rels_path = ExcelProcessor._rels_for_part(target_path)
            cell_rels = ExcelProcessor._read_relationships(zf, cell_rels_path)
            for embed_rid, img_row_idx, img_col_idx in ExcelProcessor._extract_embeds_with_location(root):
                img_target = cell_rels.get(embed_rid)
                if not img_target:
                    continue
                img_path = ExcelProcessor._resolve_target(cell_rels_path, img_target)
                if not img_path:
                    continue
                data_uri = ExcelProcessor._read_image_as_data_uri(zf, img_path)
                if not data_uri:
                    continue

                if img_row_idx is None or img_col_idx is None:
                    if row_idx is None or col_idx is None:
                        continue
                    img_row_idx, img_col_idx = row_idx, col_idx

                results.append((img_row_idx, img_col_idx, data_uri))

        else:
            # direct image target
            data_uri = ExcelProcessor._read_image_as_data_uri(zf, target_path)
            if not data_uri:
                return results
            if row_idx is None or col_idx is None:
                return results
            results.append((row_idx, col_idx, data_uri))

        return results

    @staticmethod
    def _read_relationships(zf, rels_path: str) -> Dict[str, str]:
        if not rels_path or rels_path not in zf.namelist():
            return {}
        try:
            rels_root = ET.fromstring(zf.read(rels_path))
        except Exception:
            return {}
        rels = {}
        for rel in rels_root.iter():
            if ExcelProcessor._localname(rel.tag) != 'Relationship':
                continue
            rid = rel.attrib.get('Id')
            target = rel.attrib.get('Target')
            if rid and target:
                rels[rid] = target
        return rels

    @staticmethod
    def _rels_for_part(part_path: str) -> str:
        base = posixpath.basename(part_path)
        folder = posixpath.dirname(part_path)
        return posixpath.join(folder, '_rels', f"{base}.rels")

    @staticmethod
    def _resolve_target(base_path: str, target: str) -> Optional[str]:
        if not base_path or not target:
            return None
        base_dir = posixpath.dirname(base_path)
        return posixpath.normpath(posixpath.join(base_dir, target))

    @staticmethod
    def _localname(tag: str) -> str:
        if not tag:
            return ''
        return tag.split('}', 1)[-1]

    @staticmethod
    def _find_rid(elem: ET.Element) -> Optional[str]:
        for k, v in elem.attrib.items():
            if k.endswith('id'):
                return v
        return None

    @staticmethod
    def _find_cell_location(elem: ET.Element, parent_map: Optional[Dict[ET.Element, ET.Element]] = None) -> Tuple[Optional[int], Optional[int]]:
        # 1) attribute like ref="A2" or r="A2" or f="A2"
        for k, v in elem.attrib.items():
            if k.endswith('ref') or k.endswith('cell') or k.endswith('r') or k.endswith('f'):
                loc = ExcelProcessor._cell_ref_to_pd_indices(v)
                if loc:
                    return loc

        # 2) descendant marker with row/col
        marker = ExcelProcessor._find_marker(elem)
        if marker:
            return marker

        # 3) walk up ancestors to find a ref
        if parent_map:
            cur = parent_map.get(elem)
            while cur is not None:
                for k, v in cur.attrib.items():
                    if k.endswith('ref') or k.endswith('cell') or k.endswith('r') or k.endswith('f'):
                        loc = ExcelProcessor._cell_ref_to_pd_indices(v)
                        if loc:
                            return loc
                cur = parent_map.get(cur)

        return None, None

    @staticmethod
    def _find_marker(elem: ET.Element) -> Optional[Tuple[int, int]]:
        for child in elem.iter():
            local = ExcelProcessor._localname(child.tag).lower()
            if local in ['from', 'marker']:
                row = None
                col = None
                for gc in child.iter():
                    gname = ExcelProcessor._localname(gc.tag).lower()
                    if gname == 'row':
                        try:
                            if gc.text is not None: row = int(gc.text)
                        except Exception:
                            pass
                    elif gname == 'col':
                        try:
                            if gc.text is not None: col = int(gc.text)
                        except Exception:
                            pass
                if row is not None and col is not None:
                    return row, col
        return None

    @staticmethod
    def _cell_ref_to_pd_indices(ref: str) -> Optional[Tuple[int, int]]:
        if not ref or not isinstance(ref, str):
            return None
        m = re.match(r'^([A-Z]+)([0-9]+)$', ref.strip().upper())
        if not m:
            return None
        col_str, row_str = m.groups()
        col_idx = 0
        for ch in col_str:
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)
        col_idx -= 1
        try:
            row_idx = int(row_str) - 1
        except Exception:
            return None
        return row_idx, col_idx

    @staticmethod
    def _extract_embeds_with_location(root: ET.Element) -> List[Tuple[str, Optional[int], Optional[int]]]:
        results: List[Tuple[str, Optional[int], Optional[int]]] = []
        seen = set()
        for elem in root.iter():
            embed_rid = ExcelProcessor._find_blip_embed(elem)
            if not embed_rid:
                continue
            row_idx, col_idx = ExcelProcessor._find_cell_location(elem)
            key = (embed_rid, row_idx, col_idx)
            if key in seen:
                continue
            seen.add(key)
            results.append((embed_rid, row_idx, col_idx))
        return results

    @staticmethod
    def _find_blip_embed(elem: ET.Element) -> Optional[str]:
        for child in elem.iter():
            if ExcelProcessor._localname(child.tag).lower() != 'blip':
                continue
            for k, v in child.attrib.items():
                if k.endswith('embed'):
                    return v
        return None

    @staticmethod
    def _read_image_as_data_uri(zf, image_path: str) -> Optional[str]:
        if not image_path or image_path not in zf.namelist():
            return None
        try:
            data = zf.read(image_path)
        except Exception:
            return None

        ext = os.path.splitext(image_path)[1].lower().lstrip('.')
        if not ext:
            ext = 'png'
        if ext in ['emf', 'wmf', 'vml']:
            return None
        if ext not in ['png', 'jpg', 'jpeg', 'bmp', 'gif', 'webp']:
            ext = 'png'

        b64 = base64.b64encode(data).decode('utf-8')
        return f"data:image/{ext};base64,{b64}"

    @staticmethod
    def _extract_drawing_images_map(temp_file_path: str, sheet_number: int) -> Dict[Tuple[int, int], List[str]]:
        image_map: Dict[Tuple[int, int], List[str]] = {}
        if not zipfile.is_zipfile(temp_file_path):
            return image_map

        with zipfile.ZipFile(temp_file_path, 'r') as zf:
            sheet_path = f"xl/worksheets/sheet{sheet_number}.xml"
            if sheet_path not in zf.namelist():
                return image_map

            sheet_xml = zf.read(sheet_path)
            try:
                sheet_root = ET.fromstring(sheet_xml)
            except Exception:
                return image_map

            drawing_rids = []
            for elem in sheet_root.iter():
                if ExcelProcessor._localname(elem.tag) != 'drawing':
                    continue
                rid = ExcelProcessor._find_rid(elem)
                if rid:
                    drawing_rids.append(rid)

            if not drawing_rids:
                return image_map

            sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_number}.xml.rels"
            sheet_rels = ExcelProcessor._read_relationships(zf, sheet_rels_path)

            for rid in drawing_rids:
                target = sheet_rels.get(rid)
                if not target:
                    continue
                drawing_path = ExcelProcessor._resolve_target(sheet_rels_path, target)
                if not drawing_path or drawing_path not in zf.namelist():
                    continue

                try:
                    drawing_root = ET.fromstring(zf.read(drawing_path))
                except Exception:
                    continue

                drawing_rels_path = ExcelProcessor._rels_for_part(drawing_path)
                drawing_rels = ExcelProcessor._read_relationships(zf, drawing_rels_path)

                for anchor in drawing_root.iter():
                    local = ExcelProcessor._localname(anchor.tag)
                    if local not in ['twoCellAnchor', 'oneCellAnchor']:
                        continue

                    frm = None
                    for c in anchor:
                        if ExcelProcessor._localname(c.tag) == 'from':
                            frm = c
                            break
                    if frm is None:
                        continue

                    row = None
                    col = None
                    for c in frm:
                        lname = ExcelProcessor._localname(c.tag)
                        if lname == 'row':
                            try: row = int(c.text)
                            except Exception: pass
                        elif lname == 'col':
                            try: col = int(c.text)
                            except Exception: pass
                    if row is None or col is None:
                        continue

                    blip_rid = None
                    for elem in anchor.iter():
                        if ExcelProcessor._localname(elem.tag) != 'blip':
                            continue
                        for k, v in elem.attrib.items():
                            if k.endswith('embed'):
                                blip_rid = v
                                break
                        if blip_rid:
                            break
                    if not blip_rid:
                        continue

                    img_target = drawing_rels.get(blip_rid)
                    if not img_target:
                        continue
                    img_path = ExcelProcessor._resolve_target(drawing_rels_path, img_target)
                    if not img_path:
                        continue
                    data_uri = ExcelProcessor._read_image_as_data_uri(zf, img_path)
                    if not data_uri:
                        continue

                    pd_row_idx = row - 1
                    if pd_row_idx < 0:
                        continue

                    key = (pd_row_idx, col)
                    if key not in image_map:
                        image_map[key] = []
                    image_map[key].append(data_uri)

        return image_map

    @staticmethod
    def has_embedded_images(temp_file_path: str, sheet_number: int) -> bool:
        if not temp_file_path or not zipfile.is_zipfile(temp_file_path):
            return False
        try:
            with zipfile.ZipFile(temp_file_path, 'r') as zf:
                sheet_path = f"xl/worksheets/sheet{sheet_number}.xml"
                if sheet_path not in zf.namelist():
                    return False
                xml = zf.read(sheet_path)
                if b"<drawing" in xml or b"cellImage" in xml or b"cellimage" in xml:
                    return True

                sheet_rels_path = f"xl/worksheets/_rels/sheet{sheet_number}.xml.rels"
                if sheet_rels_path in zf.namelist():
                    rels_xml = zf.read(sheet_rels_path)
                    if b"drawing" in rels_xml or b"cellimage" in rels_xml:
                        return True

                for name in zf.namelist():
                    if name.startswith('xl/media/') and name.lower().endswith(('.png','.jpg','.jpeg','.bmp','.gif','.webp')):
                        return True
        except Exception:
            return False
        return False

    @staticmethod
    def apply_sheet_updates_preserve_images(temp_file_path: str, sheet_number: int,
                                            updates: List[Tuple[int, int, str]]):
        if not updates:
            return
        if not temp_file_path or not zipfile.is_zipfile(temp_file_path):
            return

        sheet_path = f"xl/worksheets/sheet{sheet_number}.xml"

        fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(fd)

        with zipfile.ZipFile(temp_file_path, 'r') as zin:
            if sheet_path not in zin.namelist():
                return
            sheet_xml = zin.read(sheet_path)
            updated_xml = ExcelProcessor._update_sheet_xml(sheet_xml, updates)

            with zipfile.ZipFile(tmp_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == sheet_path:
                        continue
                    zout.writestr(item, zin.read(item.filename))
                zout.writestr(sheet_path, updated_xml)

        os.replace(tmp_path, temp_file_path)

    @staticmethod
    def _update_sheet_xml(sheet_xml: bytes, updates: List[Tuple[int, int, str]]) -> bytes:
        ns = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
        root = ET.fromstring(sheet_xml)
        sheet_data = root.find(f'{{{ns}}}sheetData')
        if sheet_data is None:
            sheet_data = ET.SubElement(root, f'{{{ns}}}sheetData')

        rows = {}
        for row in sheet_data.findall(f'{{{ns}}}row'):
            r = row.attrib.get('r')
            if r is not None:
                try: rows[int(r)] = row
                except Exception: pass

        for pd_row_idx, col_idx, value in updates:
            excel_row = pd_row_idx + 2
            excel_col = col_idx + 1
            cell_ref = f"{ExcelProcessor._col_to_letter(excel_col)}{excel_row}"

            row_elem = rows.get(excel_row)
            if row_elem is None:
                row_elem = ET.Element(f'{{{ns}}}row', {'r': str(excel_row)})
                inserted = False
                for i, existing in enumerate(sheet_data.findall(f'{{{ns}}}row')):
                    r = existing.attrib.get('r')
                    if r is None:
                        continue
                    try:
                        if int(r) > excel_row:
                            sheet_data.insert(i, row_elem)
                            inserted = True
                            break
                    except Exception:
                        continue
                if not inserted:
                    sheet_data.append(row_elem)
                rows[excel_row] = row_elem

            cell_elem = None
            for c in row_elem.findall(f'{{{ns}}}c'):
                if c.attrib.get('r') == cell_ref:
                    cell_elem = c
                    break

            if cell_elem is None:
                cell_elem = ET.SubElement(row_elem, f'{{{ns}}}c', {'r': cell_ref})

            for child in list(cell_elem):
                cell_elem.remove(child)

            cell_elem.set('t', 'inlineStr')
            is_elem = ET.SubElement(cell_elem, f'{{{ns}}}is')
            t_elem = ET.SubElement(is_elem, f'{{{ns}}}t')
            text = '' if value is None else str(value)
            if text.startswith(' ') or text.endswith(' '):
                t_elem.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
            t_elem.text = text

        return ET.tostring(root, encoding='utf-8', xml_declaration=True)

    @staticmethod
    def _col_to_letter(col: int) -> str:
        result = []
        while col > 0:
            col, rem = divmod(col - 1, 26)
            result.append(chr(65 + rem))
        return ''.join(reversed(result))

    @staticmethod
    def download_url_to_base64(url: str) -> Optional[str]:
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36"
        }
        try:
            # timeout防止卡死，verify=False防止内网证书报错
            response = requests.get(url, headers=headers, timeout=10, verify=False)
            if response.status_code == 200:
                content_type = response.headers.get('Content-Type', '').lower()
                
                # 宽松判断，防止服务器不返回标准 image 类型
                if 'image' in content_type or any(url.lower().endswith(x) for x in ['.jpg','.png','.jpeg','.webp']):
                    b64 = base64.b64encode(response.content).decode('utf-8')
                    
                    mime = content_type if 'image' in content_type else mimetypes.guess_type(url)[0]
                    if not mime: mime = 'image/png'
                    
                    return f"data:{mime};base64,{b64}"
        except Exception: pass
        return None

    @staticmethod
    def get_image_info(base64_str: str) -> Dict[str, str]:
        try:
            if base64_str.startswith('data:'):
                h = base64_str.split(';')[0]
                m = h.split(':')[1]
                f = m.split('/')[1]
                return {"mime_type": m, "format": f}
        except: pass
        return {"mime_type": "image/png", "format": "png"}

    @staticmethod
    def save_output_file(wb: Any, output_path: str, original_filename: Optional[str], custom_file_name: Optional[str] = None) -> Tuple[bytes, str]:
        base_name, _ = os.path.splitext(original_filename or "")
        if custom_file_name and custom_file_name.strip():
            new_filename = f"{custom_file_name.strip()}.xlsx"
        else:
            new_filename = f"smart_{base_name}.xlsx"

        if wb:
            try: wb.save(output_path)
            except: pass
            
        with open(output_path, 'rb') as f:
            data = f.read()
        return data, new_filename

    @staticmethod
    def clean_paths(paths: List[str]):
        for p in paths:
            if p and os.path.exists(p):
                try: os.remove(p)
                except: pass
    
    @staticmethod
    def parse_range(range_str: str, max_rows: int) -> Dict[str, Any]:
        parts = range_str.upper().strip().replace('：', ':').split(':')
        def parse(s):
            m = re.match(r"([A-Z]+)([0-9]+)", s)
            if not m: return 0, 0
            c_str, r_num = m.groups()
            c_idx = 0
            for char in c_str: c_idx = c_idx * 26 + (ord(char) - ord('A')) + 1
            return c_idx - 1, max(0, int(r_num) - 2)
        sc, sr = parse(parts[0])
        ec, er = parse(parts[1]) if len(parts) > 1 else (sc, max_rows - 1)
        col_match = re.match(r"([A-Z]+)", parts[0])
        col_name = col_match.group(1) if col_match else ""
        return {'col_idx': sc, 'start_row': sr, 'end_row': min(er, max_rows - 1), 'col_name': col_name}

    @staticmethod
    def get_indices_list(coord_str: str, max_rows: int) -> List[Dict]:
        return [ExcelProcessor.parse_range(c.strip(), max_rows) for c in coord_str.replace('，', ',').split(',')]

    @staticmethod
    def validate_coord_format(coord: str, is_single_col_tool: bool) -> Tuple[bool, str]:
        if not coord or not coord.strip(): return False, "Column cannot be empty."
        coord = coord.strip().upper().replace('：', ':').replace('，', ',')
        if is_single_col_tool and ',' in coord: return False, "Single column tool does not support multiple columns."
        parts = coord.split(',')
        for part in parts:
            if not re.match(r'^[A-Z]+[0-9]+(:[A-Z]+[0-9]+)?$', part.strip()): return False, f"Invalid coords: {part}"
            if is_single_col_tool and ':' in part:
                 if part.split(':')[0][0] != part.split(':')[1][0]: return False, "Single col tool cannot span columns."
        return True, ""
